import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


def _rows_and_headers(summary_rows: list):
    if not summary_rows:
        return [], []
    headers = list(summary_rows[0].keys())
    for row in summary_rows[1:]:
        for k in row.keys():
            if k not in headers:
                headers.append(k)
    rows = [[row.get(h, "") for h in headers] for row in summary_rows]
    return headers, rows


def export_to_xlsx_bytes(summary_rows: list) -> bytes:
    headers, rows = _rows_and_headers(summary_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0E6E55", end_color="0E6E55", fill_type="solid")
    for row in rows:
        ws.append(row)
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_csv_bytes(summary_rows: list) -> bytes:
    headers, rows = _rows_and_headers(summary_rows)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def export_to_pdf_bytes(summary_rows: list, title: str = "ETL Upload Summary") -> bytes:
    headers, rows = _rows_and_headers(summary_rows)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), title=title)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    table_data = [headers] + [[str(v) for v in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E6E55")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F3")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()